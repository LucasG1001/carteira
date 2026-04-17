import io
import re
import unicodedata
from datetime import datetime

import openpyxl

from src.core.exceptions import BusinessException
from src.modules.Portfolio.models.transaction_model import Transaction


class B3ParserService:
    HEADER_ALIASES = {
        "entry_side": {"entrada/saida"},
        "date": {"data"},
        "movement": {"movimentacao"},
        "product": {"produto"},
        "institution": {"instituicao"},
        "quantity": {"quantidade"},
        "unit_price": {"preco unitario"},
        "operation_value": {"valor da operacao"},
    }

    @staticmethod
    def _normalize_text(value: object) -> str:
        if value is None:
            return ""

        normalized = unicodedata.normalize("NFKD", str(value).strip())
        normalized = normalized.encode("ascii", "ignore").decode("ascii")
        normalized = re.sub(r"\s+", " ", normalized)
        return normalized.lower()

    @classmethod
    def _find_sheet(cls, workbook: openpyxl.Workbook):
        for sheet_name in workbook.sheetnames:
            if cls._normalize_text(sheet_name) == "movimentacao":
                return workbook[sheet_name]

        return workbook.worksheets[0]

    @classmethod
    def _map_headers(cls, sheet) -> dict[str, int]:
        headers: dict[str, int] = {}

        for idx, cell in enumerate(sheet[1]):
            normalized_header = cls._normalize_text(cell.value)
            if not normalized_header:
                continue

            for canonical_name, aliases in cls.HEADER_ALIASES.items():
                if normalized_header in aliases:
                    headers[canonical_name] = idx
                    break

        missing_columns = [name for name in cls.HEADER_ALIASES if name not in headers]
        if missing_columns:
            missing_labels = ", ".join(sorted(missing_columns))
            raise BusinessException(400, f"Colunas obrigatorias nao encontradas: {missing_labels}")

        return headers

    @classmethod
    def _normalize_ticker(cls, ticker: str, movement: str, product: str) -> str:
        normalized_ticker = ticker.upper().strip()
        normalized_movement = cls._normalize_text(movement)
        normalized_product = cls._normalize_text(product)

        is_subscription = "subscr" in normalized_movement or "recibo de subscricao" in normalized_product
        if is_subscription and re.match(r"^[A-Z]{4}\d{2}$", normalized_ticker) and normalized_ticker.endswith("13"):
            return f"{normalized_ticker[:-2]}11"

        return normalized_ticker

    @staticmethod
    def _safe_float(value: object) -> float:
        if value is None or str(value).strip() == "-":
            return 0.0

        if isinstance(value, str):
            value = value.replace(".", "").replace(",", ".")

        try:
            return float(value)
        except ValueError:
            return 0.0

    @staticmethod
    def parse_excel(file_content: bytes, upload_id: int, user_id: str) -> list[Transaction]:
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            sheet = B3ParserService._find_sheet(workbook)
            headers = B3ParserService._map_headers(sheet)
            transactions: list[Transaction] = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                movement = str(row[headers["movement"]] or "").strip()
                product = str(row[headers["product"]] or "").strip()
                entry_side = str(row[headers["entry_side"]] or "").strip().title()

                ticker = product.split(" - ")[0].strip()
                if not ticker:
                    continue

                ticker = B3ParserService._normalize_ticker(ticker, movement, product)

                raw_date = row[headers["date"]]
                if isinstance(raw_date, datetime):
                    date_value = raw_date.date()
                else:
                    try:
                        date_value = datetime.strptime(str(raw_date), "%d/%m/%Y").date()
                    except ValueError:
                        continue

                quantity = B3ParserService._safe_float(row[headers["quantity"]])
                unit_price = B3ParserService._safe_float(row[headers["unit_price"]])
                operation_value = B3ParserService._safe_float(row[headers["operation_value"]])

                transactions.append(
                    Transaction(
                        upload_id=upload_id,
                        user_id=user_id,
                        ticker=ticker,
                        operation_type=movement,
                        entry_side=entry_side,
                        date=date_value,
                        quantity=quantity,
                        unit_price=unit_price,
                        operation_value=operation_value,
                    )
                )

            return transactions
        except openpyxl.utils.exceptions.InvalidFileException:
            raise BusinessException(400, "Arquivo invalido ou corrompido. Envie um arquivo Excel (.xlsx).")
        except BusinessException:
            raise
        except Exception as exc:
            raise BusinessException(500, f"Erro inesperado ao processar arquivo B3: {str(exc)}")
